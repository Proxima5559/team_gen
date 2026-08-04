import io
import unicodedata
import openpyxl
from app.models.team import Team


class ExcelExporter:

    def _safe_val(self, val, default=""):
        if val is None:
            return default
        if hasattr(val, "value"):
            val = val.value
        if isinstance(val, str):
            normalized = unicodedata.normalize('NFKD', val)
            ascii_str = normalized.encode('ascii', 'ignore').decode('ascii')
            if not ascii_str.strip():
                return val.encode("utf-8", errors="replace").decode("utf-8")
            return ascii_str
        return val

    def export_bytes(self, team: Team) -> bytes:
        wb = openpyxl.Workbook()

        ws_info = wb.active
        ws_info.title = "Overview"
        ws_info.append(["Attribute", "Value"])
        ws_info.append(["Name", self._safe_val(team.name)])
        ws_info.append(["Country", self._safe_val(team.country)])
        ws_info.append(["League", self._safe_val(team.league)])
        ws_info.append(["Budget", f"${team.budget:,}"])
        
        form_val = team.formation.value if hasattr(team.formation, "value") else str(team.formation)
        ws_info.append(["Formation", self._safe_val(form_val)])
        ws_info.append(["Playing Style", self._safe_val(team.playing_style)])

        ws_players = wb.create_sheet(title="Players")
        ws_players.append([
            "Kit #", "Name", "Position", "Age", "Nationality",
            "Overall", "Potential", "Market Value ($)", "Foot", "Height (cm)", "Weight (kg)",
            "Pace", "Shooting", "Passing", "Dribbling", "Defending", "Physical", "GK",
            "Aggression", "Stamina", "Strength", "Jumping", "Heading"
        ])
        
        players = team.players or []
        print(f"--- ExcelExporter: Writing {len(players)} players to sheet ---")
        
        for p in players:
            player_row = [
                p.kit_number,
                self._safe_val(p.name),
                self._safe_val(p.position),
                p.age,
                self._safe_val(p.nationality),
                p.overall,
                p.potential,
                p.market_value,
                self._safe_val(p.preferred_foot),
                p.height_cm,
                p.weight_kg,
                p.pace,
                p.shooting,
                p.passing,
                p.dribbling,
                p.defending,
                p.physical,
                p.goalkeeping,
                p.aggression,
                p.stamina,
                p.strength,
                p.jumping,
                p.heading
            ]
            ws_players.append(player_row)

        ws_mv = wb.create_sheet(title="Manager & Stadium")
        ws_mv.append(["Category", "Attribute", "Value"])
        if team.manager:
            mgr = team.manager
            mgr_form = mgr.formation.value if hasattr(mgr.formation, "value") else str(mgr.formation)
            ws_mv.append(["Manager", "Name", self._safe_val(mgr.name)])
            ws_mv.append(["Manager", "Nationality", self._safe_val(mgr.nationality)])
            ws_mv.append(["Manager", "Preferred Formation", self._safe_val(mgr_form)])
            ws_mv.append(["Manager", "Style", self._safe_val(mgr.style)])
            
        if team.stadium:
            stad = team.stadium
            ws_mv.append(["Stadium", "Name", self._safe_val(stad.name)])
            ws_mv.append(["Stadium", "City", self._safe_val(stad.city)])
            ws_mv.append(["Stadium", "Capacity", stad.capacity])

        ws_id = wb.create_sheet(title="Identity & Culture")
        ws_id.append(["Category", "Attribute", "Value"])
        if team.identity:
            ident = team.identity
            ws_id.append(["Identity", "Founded", ident.founded])
            ws_id.append(["Identity", "Nickname", self._safe_val(ident.nickname)])
            ws_id.append(["Identity", "Club Name", self._safe_val(ident.club_name)])
            ws_id.append(["Identity", "Motto", self._safe_val(ident.motto)])
            ws_id.append(["Identity", "Mascot", self._safe_val(ident.mascot)])
            ws_id.append(["Identity", "Primary Color", self._safe_val(ident.primary_color)])
            ws_id.append(["Identity", "Secondary Color", self._safe_val(ident.secondary_color)])
            
        if team.fans:
            fans = team.fans
            ws_id.append(["Fans", "Group Name", self._safe_val(fans.supporter_name)])
            ws_id.append(["Fans", "Atmosphere", self._safe_val(fans.atmosphere)])
            ws_id.append(["Fans", "Avg Attendance", fans.average_attendance])
            ws_id.append(["Fans", "Reputation", self._safe_val(fans.reputation)])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()